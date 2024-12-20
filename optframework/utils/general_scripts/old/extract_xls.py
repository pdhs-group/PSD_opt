# -*- coding: utf-8 -*-
"""
Created on Mon Sep 13 17:04:50 2021

@author: xy0264
"""
import openpyxl

def extract_xls(xlsx_file):
    
    # Load ative datasheet from xlsx file 
    sheet = openpyxl.load_workbook(xlsx_file).active
    
    # Create data dictionary and maximum column "counter"
    data = {}
    maxcol = 0

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        
        if i == 0:
            # Fill keys and index the dictionary
            keys=[]
            for c in range(sheet.max_column):
                if not row[c] is None:
                    keys+=[row[c]]
                    data[row[c]] = []
                
            maxcol=len(keys)
                        
        # Fill the dictionary with corresponding values
        else:
            for c in range(maxcol):
                data[keys[c]].append(row[c])
                #print(c,row[c])
    
    #print(maxcol)
    return data